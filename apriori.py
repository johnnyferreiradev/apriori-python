import pandas as pd
import openpyxl
from mlxtend.frequent_patterns import apriori, association_rules

# loading dataset ger alta
wookbook_ger_alta = openpyxl.load_workbook("association_rules_ger/dataset_ger_alta.xlsx")
worksheet_ger_alta = wookbook_ger_alta.active

columns_ger_alta = []
dataset_ger_alta = []

for i in range(0, worksheet_ger_alta.max_row):
    row_values = []
    for col in worksheet_ger_alta.iter_cols(1, worksheet_ger_alta.max_column):
        if i == 0:
            columns_ger_alta.append(col[i].value)
        if i > 0:
            row_values.append(col[i].value)
    if i > 0:
        dataset_ger_alta.append(row_values)

df_ger_alta = pd.DataFrame(dataset_ger_alta, columns=columns_ger_alta)

# loading dataset ger baixa
wookbook_ger_baixa = openpyxl.load_workbook("association_rules_ger/dataset_ger_baixa.xlsx")
worksheet_ger_baixa = wookbook_ger_baixa.active

columns_ger_baixa = []
dataset_ger_baixa = []

for i in range(0, worksheet_ger_baixa.max_row):
    row_values = []
    for col in worksheet_ger_baixa.iter_cols(1, worksheet_ger_baixa.max_column):
        if i == 0:
            columns_ger_baixa.append(col[i].value)
        if i > 0:
            row_values.append(col[i].value)
    if i > 0:
        dataset_ger_baixa.append(row_values)

df_ger_baixa = pd.DataFrame(dataset_ger_baixa, columns=columns_ger_baixa)

class Apriori:
    threshold = 0.4
    df = None

    def __init__(self, df, threshold=None, transform_bol=False):
        self._validate_df(df)

        self.df = df
        if threshold is not None:
            self.threshold = threshold

        if transform_bol:
            self._transform_bol()


    def _validate_df(self, df=None):
        if df is None:
            raise Exception("df must be a valid pandas.DataDrame.")


    def _transform_bol(self):
        for column in self.df.columns:
            self.df[column] = self.df[column].apply(lambda x: True if x == 1 else False)


    def _apriori(self, use_colnames=False, max_len=None, count=True):
        apriori_df = apriori(
                    self.df, 
                    min_support=self.threshold,
                    use_colnames=use_colnames, 
                    max_len=max_len,
                )
        if count:
            apriori_df['length'] = apriori_df['itemsets'].apply(lambda x: len(x))

        return apriori_df

    def run(self, use_colnames=False, max_len=None, count=True):
        return self._apriori(
                        use_colnames=use_colnames,
                        max_len=max_len,
                        count=count
                    )

    def filter(self, apriori_df, length, threshold):
        if 'length' not in apriori_df.columns:
            raise Exception("apriori_df has no length. Please run the Apriori with count=True.")

        return apriori_df[ (apriori_df['length'] == length) & (apriori_df['support'] >= threshold) ]


# ger alta
if 'ID' in df_ger_alta.columns: del df_ger_alta['ID']

apriori_runner = Apriori(df_ger_alta, threshold=0.1, transform_bol=True)
apriori_df_ger_alta = apriori_runner.run(use_colnames=True)
print(apriori_df_ger_alta)
apriori_df_ger_alta.to_excel('apriori_df_ger_alta.xlsx')

association_rules_result = association_rules(apriori_df_ger_alta, metric='confidence', min_threshold=0.3, support_only=False)
print(' ')
print(association_rules_result)
association_rules_result.to_excel('association_rules_result_ger_alta.xlsx')

# ger baixa
if 'ID' in df_ger_baixa.columns: del df_ger_baixa['ID']

apriori_runner = Apriori(df_ger_baixa, threshold=0.1, transform_bol=True)
apriori_df_ger_baixa = apriori_runner.run(use_colnames=True)
print(apriori_df_ger_baixa)
apriori_df_ger_baixa.to_excel('apriori_df_ger_baixa.xlsx')

association_rules_result = association_rules(apriori_df_ger_baixa, metric='confidence', min_threshold=0.3, support_only=False)
print(' ')
print(association_rules_result)
association_rules_result.to_excel('association_rules_result_ger_baixa.xlsx')